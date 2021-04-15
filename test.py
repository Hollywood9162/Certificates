import pandas as pd
from docx import Document
from docx2pdf import convert
import openpyxl

# This script runs on Windows only, and you must have Word installed.
''' scope of work
input: excel spreadsheet with names
output: generate a certificate for each of them
'''

'''
 concepts learned: 
 * python basics: import a module, module requirements.txt, for loop and break, if condition, variable assignment, print, string object
 * pandas basics: dataframe, read_excel with parameters, iteration/loop
 * python-docx: document, paragraph, runs
 * what could your teacher tell you to improve

https://wiki.python.org/moin/BeginnersGuide/Programmers

'''

# read input excel
import openpyxl
wb = openpyxl.load_workbook('Special Topics Sheet.xlsx')
wb.sheetnames # The workbook's sheets' names.
['Music', 'Science Writing', 'Energy', 'Ancient Greek', 'Biotech Journal', 'PUBLIC HEALTH HISPANIC', 'Investing', 'MIDDLE EAST', 'Modern Healthcare']
sheet = wb['Music'] # Get a sheet from the workbook.
sheet
<Worksheet "Music">
type(sheet)
<class 'openpyxl.worksheet.worksheet.Worksheet'>
sheet.title # Get the sheet's title as a string.
'Music'
anotherSheet = wb.active # Get the active sheet.
anotherSheet
<Worksheet "Science Writing">

import openpyxl
wb = openpyxl.load_workbook('Special Topics Sheet.xlsx')
sheet = wb['Music'] # Get a sheet from the workbook.
sheet['D1,2,3,4,5,6,7,8'] # Get a cell from the sheet.
<Cell 'Music'.D2,3,4,5,6,7,8>
sheet['D1,2,3,4,5,6,7,8'].value # Get the value from the cell.
studentsname.studentsname(2015, 4, 5, 13, 34, 2)
c = sheet['L1'] # Get another cell from the sheet.
c.value
'Apples'
# Get the row, column, and value from the cell.
'Row %s, Column %s is %s' % (c.row, c.column, c.value)
'Row 1, Column B is Apples'
'Cell %s is %s' % (c.coordinate, c.value)
'Cell B1 is Apples'
sheet['C1'].value
73

import openpyxl
wb = openpyxl.load_workbook('Special Topics Sheet.xlsx')
sheet = wb['Music']
tuple(sheet['A1':'C3']) # Get all cells from A1 to C3.
((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell
   'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>,
   <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
➊ for rowOfCellObjects in sheet['A1':'C3']:
➋ ...     for cellObj in rowOfCellObjects:
   ...         print(cellObj.coordinate, cellObj.value)
   ...     print('--- END OF ROW ---')

# iterate through the rows
for row_nb, row in df.iterrows():
    print(row_nb, row)
    student_name = row["Student's Name"]
    print(student_name)

    # open document template
    document = Document("Certificate of Achievement Template.docx")

    # go through all the paragraphs
    for paragraph in document.paragraphs:
        # and all the "runs" in each paragraph
        for run in paragraph.runs:
            # if the run contains Name
            if 'NAME' in run.text:
                # replace with student_name
                run.text = student_name
                break
    document.save(f'certificates/certificate_{student_name}.docx')

    #convert docx to pdf
    convert(f'certificates/certificate_{student_name}.docx')
    convert("input.docx", "output.pdf")
    convert("certificates/")

    document.save(f"certificates/certificate_{student_name}.pdf")

    # for each student_name, generate a new word doc/certificate
    ## open the template "Certificate of Achievement Template.docx"
    ## search/replace the text "Name" with student_name
    ## save the docx to a new file with format like certificate_student_name.docx
    ## save the docx as PDF

